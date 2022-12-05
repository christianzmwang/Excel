
from openpyxl import load_workbook, Workbook

wb = Workbook()
ws = wb.create_sheet('eg')

data = [('id', 'name', 'country'),
        (1, 'Trudeau','Canada'),
        (2,'Zelenskyy', 'Ukraine'),
        (3,'Putin', 'Russia'),
        (4,),
        (5,'Biden', 'US'),
        (6, 'Xi', 'China'),
        (7,'Johnson', 'UK'),
        (8, 'Castex', 'France'),
        (9,),
        (10, 'Steinmeier', 'Germany'),
        (11, 'Rutte', 'Netherlands'),
        (12,'Loong', 'Singapore')]

for row in data:
  ws.append(row)

wb.save("names.xlsx")


